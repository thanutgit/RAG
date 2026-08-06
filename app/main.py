"""
app/main.py
FastAPI application — ห่อ ingest_service และ query_service เป็น REST API

รัน: uvicorn app.main:app --reload --port 8000
ดู API docs อัตโนมัติที่: http://localhost:8000/docs
"""

import io
import uuid
from pathlib import Path

from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from services import config, postgres_service, qdrant_service
from services.ingest_service import run_ingest
from services.query_service import run_query

app = FastAPI(
    title="Obsidian RAG API",
    description="ระบบถาม-ตอบจากโน้ต Obsidian ผ่าน RAG (รันบน local 100%)",
    version="0.1.0",
)


# ---------------------------------------------------------------- schemas
# Pydantic model = ตัวตรวจสอบรูปแบบข้อมูลที่ client ส่งเข้ามา
# ถ้า client ส่งข้อมูลผิด type หรือขาด field ที่ required, FastAPI จะตอบ 422 ให้อัตโนมัติ


class IngestRequest(BaseModel):
    force: bool = Field(
        default=False,
        description="บังคับ re-ingest ทุกไฟล์ ไม่สนใจว่าเปลี่ยนแปลงหรือไม่",
    )


class QueryRequest(BaseModel):
    question: str = Field(..., min_length=1, description="คำถามที่จะถาม")
    top_k: int | None = Field(default=None, ge=1, le=20, description="จำนวน chunk สูงสุดที่จะดึงมาอ้างอิง")
    score_threshold: float | None = Field(
        default=None,
        ge=0,
        le=1,
        description="ตัด chunk ที่ score ต่ำกว่านี้ทิ้ง (ค่า default มาจาก .env)",
    )
    session_id: str | None = Field(default=None, description="ถ้าไม่ส่งมา จะสร้าง session ใหม่ให้")


class Source(BaseModel):
    file_path: str
    chunk_index: int
    heading: str = ""
    score: float
    text_preview: str


class ResultTable(BaseModel):
    columns: list[str]
    rows: list[list[str | None]]
    total_rows: int
    truncated: bool


class QueryResponse(BaseModel):
    answer: str
    sources: list[Source]
    latency_ms: int
    session_id: str
    # mode บอกว่าคำตอบมาจากเส้นทางไหน — "search" (vector) หรือ "sql" (คำนวณจริง)
    mode: str = "search"
    sql: str | None = None
    # ตารางผลลัพธ์เต็ม ส่งตรงให้ frontend แสดง ไม่ผ่าน LLM
    table: ResultTable | None = None
    # ถ้าคำถามถูกเขียนใหม่จากบริบทสนทนา จะแสดงคำถามที่ใช้ค้นหาจริงให้ผู้ใช้เห็น
    rewritten_question: str | None = None


# ---------------------------------------------------------------- endpoints


@app.get("/health")
def health_check():
    """เช็คว่า service พร้อมใช้งาน และเชื่อมต่อ Qdrant ได้จริง"""
    try:
        client = qdrant_service.get_client()
        chunk_count = qdrant_service.count(client)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"เชื่อมต่อ Qdrant ไม่ได้: {e}") from e

    return {"status": "ok", "chunks_in_db": chunk_count}


@app.post("/ingest")
def ingest_endpoint(req: IngestRequest):
    """
    สั่งให้ระบบอ่าน vault ใหม่ทั้งหมด แล้ว sync เข้า Qdrant
    (ไฟล์ที่ไม่เปลี่ยนแปลงจะถูกข้าม ยกเว้นตั้ง force=true)
    """
    try:
        stats = run_ingest(force=req.force)
    except FileNotFoundError as e:
        raise HTTPException(status_code=500, detail=str(e)) from e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ingest ล้มเหลว: {e}") from e

    return stats


@app.get("/tables")
def list_tables_endpoint():
    """รายชื่อตารางที่โหลดเข้า SQL แล้ว ใช้ตรวจว่าไฟล์ตารางถูกโหลดครบไหม"""
    from services import tabular_service

    try:
        conn = postgres_service.get_connection()
        try:
            return {"tables": tabular_service.list_tables(conn)}
        finally:
            conn.close()
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"เชื่อมต่อฐานข้อมูลไม่ได้: {e}") from e


@app.post("/query", response_model=QueryResponse)
def query_endpoint(req: QueryRequest):
    """รับคำถาม -> ค้นหาจาก Qdrant -> ให้ LLM ตอบ พร้อม log ประวัติลง PostgreSQL"""
    session_id = req.session_id or str(uuid.uuid4())

    try:
        result = run_query(
            req.question,
            top_k=req.top_k,
            score_threshold=req.score_threshold,
            session_id=session_id,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Query ล้มเหลว: {e}") from e

    # log ประวัติแชท (best-effort — ถ้า log พลาดไม่ควรทำให้ user ไม่ได้คำตอบ)
    try:
        conn = postgres_service.get_connection()
        postgres_service.ensure_session(conn, session_id)
        postgres_service.log_chat_message(conn, session_id, "user", req.question)
        postgres_service.log_chat_message(
            conn,
            session_id,
            "assistant",
            result["answer"],
            model_name=config.OLLAMA_LLM_MODEL,
            retrieved_chunks=result["sources"],
            latency_ms=result["latency_ms"],
        )
        conn.close()
    except Exception:
        pass  # ไม่ให้การ log พังทำให้ทั้ง request fail

    return {**result, "session_id": session_id}


class ExportRequest(BaseModel):
    columns: list[str] = Field(..., min_length=1)
    rows: list[list[str | None]]
    filename: str = Field(default="query_result", max_length=80)


@app.post("/export/xlsx")
def export_xlsx(req: ExportRequest):
    """
    แปลงผลลัพธ์เป็นไฟล์ Excel ให้ดาวน์โหลด

    ทำฝั่ง server แทนที่จะให้ browser สร้าง CSV เอง เพราะ:
      - Excel เดา encoding ของ CSV ผิดบ่อย ทำให้ภาษาไทยเพี้ยน
      - ค่าอย่าง "2026-01" ถูก Excel แปลงเป็นวันที่โดยไม่ได้ตั้งใจ
      - ไฟล์ xlsx จริงเก็บชนิดข้อมูลไว้ได้ ตัวเลขยังเป็นตัวเลข
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="ต้องติดตั้ง openpyxl ก่อน") from None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ผลลัพธ์"

    ws.append(req.columns)
    header_fill = PatternFill("solid", fgColor="E9EBE4")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for row in req.rows:
        # แปลงกลับเป็นตัวเลขถ้าค่านั้นเป็นตัวเลขจริง เพื่อให้ Excel คำนวณต่อได้
        converted = []
        for v in row:
            if v is None or v == "":
                converted.append(None)
                continue
            try:
                converted.append(int(v))
            except (ValueError, TypeError):
                try:
                    converted.append(float(v))
                except (ValueError, TypeError):
                    converted.append(v)
        ws.append(converted)

    # ปรับความกว้างคอลัมน์ให้พออ่าน โดยดูจากค่าที่ยาวที่สุด (จำกัดที่ 50)
    for i, col in enumerate(req.columns, start=1):
        longest = max(
            [len(str(col))] + [len(str(r[i - 1])) for r in req.rows[:200] if i - 1 < len(r) and r[i - 1]]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 3, 10), 50)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c for c in req.filename if c.isalnum() or c in "-_")[:60] or "result"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )


# ---------------------------------------------------------------- frontend
# เสิร์ฟหน้าเว็บจาก origin เดียวกับ API เพื่อไม่ต้องตั้ง CORS
# ต้อง mount ท้ายสุด เพราะ path "/" จะจับทุก request ที่ไม่ตรง endpoint ด้านบน
_frontend_dir = Path(__file__).resolve().parent.parent / "frontend"
if _frontend_dir.exists():
    app.mount("/", StaticFiles(directory=_frontend_dir, html=True), name="frontend")
