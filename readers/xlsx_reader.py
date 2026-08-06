"""
readers/xlsx_reader.py
อ่าน .xlsx / .xlsm แปลงเป็น Markdown

ต่างจาก CSV ตรงที่ spreadsheet มีความซับซ้อนเพิ่มหลายชั้น:
  1. หลายชีท          -> แยกเป็นคนละ section (## ต่อชีท) ไม่ปนกัน
  2. สูตร              -> อ่าน "ค่าที่คำนวณแล้ว" ไม่ใช่ตัวสูตร เพราะ =SUMIFS(...) ไม่มีประโยชน์ตอนค้นหา
  3. โครงสร้างไม่แน่นอน -> บางชีทเป็นตาราง บางชีทเป็น key-value หรือข้อความอธิบาย
  4. header ไม่ได้อยู่แถวแรกเสมอ -> มักมีหัวเรื่อง/คำอธิบายก่อน
  5. กราฟ/รูป          -> อ่านไม่ได้ แต่บันทึกไว้ว่ามีอยู่

ข้อจำกัดของ data_only=True:
  openpyxl อ่านค่าที่ Excel cache ไว้ ถ้าไฟล์ถูกสร้างด้วยโปรแกรมและไม่เคยเปิดใน
  Excel/LibreOffice มาก่อน จะไม่มีค่า cache -> ได้ None
  กรณีนั้นระบบจะ fallback ไปอ่านสูตรแทน พร้อมแจ้งเตือน
"""

from pathlib import Path

from readers.base import Document, ReaderError, escape_md_cell, resolve_path

ROWS_PER_BLOCK = 25
MAX_ROWS_PER_SHEET = 2000
MAX_SCAN_FOR_HEADER = 10  # แถวแรก ๆ ที่จะสแกนหา header


def _fmt(value) -> str:
    """แปลงค่าในเซลล์เป็นข้อความที่อ่านง่าย"""
    if value is None:
        return ""
    if isinstance(value, float):
        # เลขจำนวนเต็มไม่ต้องมี .0 / ทศนิยมปัดเหลือ 4 ตำแหน่ง
        if value.is_integer():
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _rows_of(ws, max_rows: int) -> list[list[str]]:
    """ดึงแถวที่มีข้อมูลจริง ตัดคอลัมน์ว่างท้ายแถวออก"""
    rows = []
    for row in ws.iter_rows(max_row=min(ws.max_row, max_rows), values_only=True):
        cells = [_fmt(c) for c in row]
        while cells and not cells[-1]:
            cells.pop()
        if any(cells):
            rows.append(cells)
    return rows


def _find_header(rows: list[list[str]]) -> int | None:
    """
    หาแถวที่น่าจะเป็น header ของตาราง

    เกณฑ์: แถวที่มีหลายคอลัมน์ (>=3) เต็มทุกช่อง ไม่มีตัวเลขล้วน และแถวถัดไป
    มีจำนวนคอลัมน์ใกล้เคียงกัน -> น่าจะเป็นหัวตารางที่มีข้อมูลตามมา
    """
    for i, row in enumerate(rows[:MAX_SCAN_FOR_HEADER]):
        if len(row) < 3 or not all(c for c in row):
            continue
        if all(c.replace(".", "").replace("-", "").isdigit() for c in row):
            continue
        following = rows[i + 1 : i + 4]
        if following and sum(len(r) >= len(row) - 1 for r in following) >= min(2, len(following)):
            return i
    return None


def _is_section_break(row: list[str], prev_width: int) -> bool:
    """
    แถวนี้เป็นหัวข้อใหม่ที่คั่นระหว่างตารางหรือไม่

    สัญญาณ: มีข้อมูลช่องเดียว (หัวข้อลอย ๆ) ในขณะที่แถวก่อนหน้ากว้างกว่ามาก
    เช่น "📅 สรุปรายเดือนทั้งปี" ที่คั่นระหว่างตารางหมวดหมู่กับตารางรายเดือน
    """
    filled = [c for c in row if c]
    return len(filled) == 1 and prev_width >= 3


def _split_blocks(rows: list[list[str]]) -> list[list[list[str]]]:
    """แบ่งชีทเป็นบล็อกย่อยเมื่อเจอแถวหัวข้อ เพื่อไม่ให้ตารางคนละชุดถูกรวมเป็นอันเดียว"""
    blocks: list[list[list[str]]] = []
    current: list[list[str]] = []
    prev_width = 0

    for row in rows:
        if current and _is_section_break(row, prev_width):
            blocks.append(current)
            current = [row]
        else:
            current.append(row)
        prev_width = len([c for c in row if c])

    if current:
        blocks.append(current)
    return blocks


def _sheet_as_table(rows: list[list[str]], header_idx: int) -> list[str]:
    """แปลงเป็นตาราง Markdown โดยแนบ header ซ้ำทุกบล็อก"""
    parts = []
    preamble = rows[:header_idx]
    header = rows[header_idx]
    data = rows[header_idx + 1 :]

    for r in preamble:
        text = " ".join(c for c in r if c)
        if text:
            parts.extend([text, ""])

    header_cells = [escape_md_cell(c) or f"คอลัมน์{i + 1}" for i, c in enumerate(header)]
    header_line = "| " + " | ".join(header_cells) + " |"
    sep_line = "| " + " | ".join("---" for _ in header_cells) + " |"

    parts.append(f"ตาราง {len(data)} แถว คอลัมน์: {', '.join(header_cells)}")
    parts.append("")

    for start in range(0, len(data), ROWS_PER_BLOCK):
        block = data[start : start + ROWS_PER_BLOCK]
        if len(data) > ROWS_PER_BLOCK:
            parts.append(f"### แถวที่ {start + 1}-{start + len(block)}")
            parts.append("")
        parts.append(header_line)
        parts.append(sep_line)
        for row in block:
            cells = [escape_md_cell(row[i]) if i < len(row) else "" for i in range(len(header_cells))]
            parts.append("| " + " | ".join(cells) + " |")
        parts.append("")

    return parts


def _sheet_as_text(rows: list[list[str]]) -> list[str]:
    """
    ชีทที่ไม่ใช่ตาราง (เช่น dashboard หรือคู่มือ) -> แปลงเป็นข้อความบรรทัดต่อบรรทัด
    2 คอลัมน์ที่มีข้อมูลจะถูกมองเป็นคู่ key-value ซึ่งอ่านเข้าใจง่ายกว่าตาราง
    """
    parts = []
    for row in rows:
        filled = [c for c in row if c]
        if not filled:
            continue
        if len(filled) == 1:
            parts.append(filled[0])
        elif len(filled) == 2:
            parts.append(f"**{filled[0]}**: {filled[1]}")
        else:
            parts.append(" · ".join(filled))
        parts.append("")
    return parts


def read(path: str | Path) -> Document:
    p = resolve_path(path)

    try:
        import openpyxl
    except ImportError:
        raise ReaderError("ต้องติดตั้ง openpyxl ก่อน: pip install openpyxl") from None

    try:
        wb_values = openpyxl.load_workbook(str(p), data_only=True, read_only=False)
    except Exception as e:
        raise ReaderError(f"เปิดไฟล์ไม่ได้: {p.name} ({e})") from e

    parts = [f"# {p.stem}", ""]
    meta = {"sheets": [], "charts": 0, "formulas_uncached": False}

    for ws in wb_values.worksheets:
        if ws.sheet_state != "visible":
            continue

        rows = _rows_of(ws, MAX_ROWS_PER_SHEET)
        if not rows:
            continue

        parts.append(f"## {ws.title}")
        parts.append("")

        # ชีทเดียวอาจมีหลายตาราง คั่นด้วยแถวหัวข้อ -> แยกเป็นบล็อกก่อน
        for block in _split_blocks(rows):
            header_idx = _find_header(block)
            if header_idx is not None:
                parts.extend(_sheet_as_table(block, header_idx))
            else:
                parts.extend(_sheet_as_text(block))

        n_charts = len(getattr(ws, "_charts", []) or [])
        n_images = len(getattr(ws, "_images", []) or [])
        if n_charts or n_images:
            note = []
            if n_charts:
                note.append(f"กราฟ {n_charts} รายการ")
            if n_images:
                note.append(f"รูปภาพ {n_images} รายการ")
            # บันทึกไว้ว่ามีอยู่ เพราะข้อมูลในกราฟดึงเป็นข้อความไม่ได้
            parts.append(f"*(ชีทนี้มี{' และ'.join(note)} ซึ่งไม่สามารถอ่านเป็นข้อความได้)*")
            parts.append("")
            meta["charts"] += n_charts

        meta["sheets"].append(
            {
                "name": ws.title,
                "rows": len(rows),
                "as_table": header_idx is not None,
            }
        )

    text = "\n".join(parts).strip()

    # ตรวจว่าเจอสูตรที่ไม่มีค่า cache หรือไม่
    if "=" in text and any(f in text for f in ("SUM(", "SUMIFS(", "IFERROR(", "VLOOKUP(")):
        meta["formulas_uncached"] = True
        print(
            f"   ⚠️  {p.name}: พบสูตรที่ยังไม่มีค่าคำนวณ cache ไว้\n"
            f"       เปิดไฟล์ใน Excel หรือ LibreOffice แล้วบันทึกซ้ำหนึ่งครั้ง "
            f"เพื่อให้ระบบอ่านค่าจริงแทนตัวสูตร"
        )

    doc = Document(
        text=text,
        source_path=str(p),
        file_type=p.suffix.lower().lstrip("."),
        metadata=meta,
    )

    if doc.is_empty:
        raise ReaderError(f"ไม่มีข้อมูลในไฟล์: {p.name}")

    return doc


def extract_tables(path: str | Path) -> list[dict]:
    """
    คืนข้อมูลดิบสำหรับโหลดเข้า SQL (ต่างจาก read() ที่คืน Markdown)

    Excel อาจมีหลายชีท และแต่ละชีทอาจมีหลายตาราง จึงคืนได้หลายรายการ
    เอาเฉพาะบล็อกที่ตรวจแล้วว่าเป็นตารางจริง (มี header ชัดเจน) เท่านั้น
    """
    p = resolve_path(path)

    try:
        import openpyxl
    except ImportError:
        raise ReaderError("ต้องติดตั้ง openpyxl ก่อน: pip install openpyxl") from None

    wb = openpyxl.load_workbook(str(p), data_only=True)
    tables = []

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue

        rows = _rows_of(ws, MAX_ROWS_PER_SHEET)
        if not rows:
            continue

        for bi, block in enumerate(_split_blocks(rows)):
            header_idx = _find_header(block)
            if header_idx is None:
                continue

            header = [str(c).strip() for c in block[header_idx]]
            data = block[header_idx + 1 :]
            if not data:
                continue

            # ชื่อตาราง: ใช้ชื่อชีท ถ้าชีทเดียวมีหลายตารางค่อยเติมเลขต่อท้าย
            name = ws.title if bi == 0 else f"{ws.title}_{bi + 1}"
            tables.append({"name": name, "header": header, "rows": data})

    return tables
